#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel批量合并工具 - 合并多个Excel文件为一个"""

import pandas as pd
import os
import sys
from pathlib import Path


def merge_excel_files(input_dir, output_file, file_pattern="*.xlsx", sheet_name=0, use_pandas=True):
    """
    合并指定目录下所有Excel文件
    
    Args:
        input_dir: 输入目录路径
        output_file: 输出文件路径
        file_pattern: 文件匹配模式，如"*.xlsx"或"*.xls"
        sheet_name: 工作表名称或索引
        use_pandas: 使用pandas（更强大的数据合并）
    """
    input_path = Path(input_dir)
    if not input_path.exists():
        return False, f"目录不存在: {input_dir}"
    
    excel_files = list(input_path.glob(file_pattern))
    if not excel_files:
        return False, f"在 {input_dir} 中未找到匹配 {file_pattern} 的文件"
    
    excel_files.sort()
    
    try:
        if use_pandas:
            dfs = []
            for f in excel_files:
                df = pd.read_excel(f, sheet_name=sheet_name)
                df['_来源文件'] = f.name
                dfs.append(df)
            merged = pd.concat(dfs, ignore_index=True)
            merged.to_excel(output_file, index=False)
        else:
            from openpyxl import Workbook, load_workbook
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.title = "合并数据"
            first = True
            for f in excel_files:
                wb_in = load_workbook(f)
                ws_in = wb_in.active
                for row in ws_in.iter_rows(min_row=1 if first else 2, values_only=True):
                    ws_out.append(row)
                first = False
            wb_out.save(output_file)
        
        return True, f"成功合并 {len(excel_files)} 个文件到 {output_file}"
    except Exception as e:
        return False, f"合并失败: {str(e)}"


def main():
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    
    root = tk.Tk()
    root.title("Excel批量合并工具")
    root.geometry("600x450")
    root.resizable(False, False)
    
    tk.Label(root, text="Excel批量合并工具", font=("", 16, "bold")).pack(pady=10)
    
    frame = tk.Frame(root, padx=20)
    frame.pack(fill=tk.BOTH, expand=True)
    
    # Input directory
    tk.Label(frame, text="选择Excel文件所在目录:").pack(anchor=tk.W, pady=(5,0))
    dir_frame = tk.Frame(frame)
    dir_frame.pack(fill=tk.X, pady=5)
    dir_var = tk.StringVar()
    tk.Entry(dir_frame, textvariable=dir_var).pack(side=tk.LEFT, fill=tk.X, expand=True)
    tk.Button(dir_frame, text="浏览...", command=lambda: dir_var.set(filedialog.askdirectory())).pack(side=tk.RIGHT, padx=(5,0))
    
    # File pattern
    tk.Label(frame, text="文件匹配模式:").pack(anchor=tk.W, pady=(10,0))
    pattern_var = tk.StringVar(value="*.xlsx")
    tk.Entry(frame, textvariable=pattern_var).pack(fill=tk.X, pady=5)
    
    # Output file
    tk.Label(frame, text="选择输出文件:").pack(anchor=tk.W, pady=(10,0))
    out_frame = tk.Frame(frame)
    out_frame.pack(fill=tk.X, pady=5)
    out_var = tk.StringVar(value="合并结果.xlsx")
    tk.Entry(out_frame, textvariable=out_var).pack(side=tk.LEFT, fill=tk.X, expand=True)
    tk.Button(out_frame, text="选择...", command=lambda: out_var.set(filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel文件","*.xlsx")]))).pack(side=tk.RIGHT, padx=(5,0))
    
    # Log
    tk.Label(frame, text="运行日志:").pack(anchor=tk.W, pady=(10,0))
    log_text = tk.Text(frame, height=8, state=tk.DISABLED)
    log_text.pack(fill=tk.BOTH, expand=True, pady=5)
    
    def log(msg):
        log_text.config(state=tk.NORMAL)
        log_text.insert(tk.END, msg + "\n")
        log_text.see(tk.END)
        log_text.config(state=tk.DISABLED)
        root.update()
    
    def run():
        input_dir = dir_var.get().strip()
        pattern = pattern_var.get().strip()
        output_file = out_var.get().strip()
        
        if not input_dir or not output_file:
            messagebox.showerror("错误", "请填写所有字段")
            return
        
        log("正在合并...")
        success, msg = merge_excel_files(input_dir, output_file, pattern)
        log(msg)
        if success:
            messagebox.showinfo("完成", msg)
    
    tk.Button(frame, text="开始合并", command=run, bg="#4CAF50", fg="white", font=("", 12, "bold")).pack(pady=10)
    
    root.mainloop()


if __name__ == "__main__":
    if len(sys.argv) > 1:
        success, msg = merge_excel_files(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "合并结果.xlsx", sys.argv[3] if len(sys.argv) > 3 else "*.xlsx")
        print(msg)
        sys.exit(0 if success else 1)
    else:
        main()
